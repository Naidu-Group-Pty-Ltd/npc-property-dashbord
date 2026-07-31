"""Reusable Excel building blocks for the Aurixa Finance Portal templates.

Mirrors ``docx_kit`` so the workbook and the Word agreements share one visual
language: obsidian banners, gold section rules, sand label cells and pale-gold
input cells. Every helper takes colours from ``aurixa_brand`` rather than
hard-coding hex, so a partner re-skin is a one-line change.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

from aurixa_brand import PALETTE, BrandProfile

# Excel wants ARGB; the palette stores RGB.
def argb(hex_rgb: str) -> str:
    return f"FF{hex_rgb}"


FONT_BODY = "Calibri"
FONT_DISPLAY = "Georgia"

# Number formats
MONEY = '"$"#,##0'
MONEY_CENTS = '"$"#,##0.00'
PERCENT = "0.00%"
DATE = "dd/mm/yyyy"

NO_FILL = PatternFill()


def fill(hex_rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb(hex_rgb))


def side(hex_rgb: str, style: str = "thin") -> Side:
    return Side(style=style, color=argb(hex_rgb))


HAIRLINE = Border(bottom=side(PALETTE.line))
FIELD_BORDER = Border(bottom=side(PALETTE.gold_tint), top=side(PALETTE.paper_warm))
CELL_BOX = Border(
    left=side(PALETTE.line), right=side(PALETTE.line),
    top=side(PALETTE.line), bottom=side(PALETTE.line),
)

WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


# --------------------------------------------------------------------------
# Blocks
# --------------------------------------------------------------------------

def banner(ws, cell_range: str, text, brand: BrandProfile, *, size: int = 18,
           height: float = 26, subtitle: str | None = None) -> None:
    """Obsidian title banner with a gold underline — the sheet's masthead."""
    ws.merge_cells(cell_range)
    first = cell_range.split(":")[0]
    cell = ws[first]
    cell.value = text
    cell.font = Font(name=FONT_DISPLAY, size=size, bold=True, color=argb(PALETTE.paper))
    cell.fill = fill(brand.primary)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    start_row = int("".join(c for c in first if c.isdigit()))
    end_row = int("".join(c for c in cell_range.split(":")[1] if c.isdigit()))
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = height
        for col in range(1, _last_col(cell_range) + 1):
            target = ws.cell(row, col)
            if target.fill.patternType is None:
                target.fill = fill(brand.primary)
            if row == end_row:
                target.border = Border(bottom=side(brand.accent, "medium"))
    if subtitle:
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1,
                                   wrap_text=True)


def _last_col(cell_range: str) -> int:
    end = cell_range.split(":")[1]
    letters = "".join(c for c in end if c.isalpha())
    total = 0
    for char in letters:
        total = total * 26 + (ord(char.upper()) - 64)
    return total


def section_bar(ws, row: int, last_col: int, text: str, brand: BrandProfile,
                *, kicker: str = "") -> None:
    """Gold section rule spanning the sheet width."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row, 1)
    cell.value = f"{text}      {kicker}" if kicker else text
    cell.font = Font(name=FONT_BODY, size=10, bold=True, color=argb(PALETTE.obsidian))
    cell.fill = fill(brand.accent)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 19
    for col in range(1, last_col + 1):
        ws.cell(row, col).fill = fill(brand.accent)


def subhead(ws, row: int, first_col: int, last_col: int, text: str,
            brand: BrandProfile) -> None:
    """Column-group heading inside a section (e.g. 'Applicant 1')."""
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    cell = ws.cell(row, first_col)
    cell.value = text
    cell.font = Font(name=FONT_BODY, size=9, bold=True, color=argb(PALETTE.gold_dark))
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    for col in range(first_col, last_col + 1):
        ws.cell(row, col).fill = fill(PALETTE.gold_tint)
        ws.cell(row, col).border = Border(bottom=side(brand.accent))


def label(ws, row: int, col: int, text: str, *, span: int = 1) -> None:
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row,
                       end_column=col + span - 1)
    cell = ws.cell(row, col)
    cell.value = text
    cell.font = Font(name=FONT_BODY, size=9, bold=True, color=argb(PALETTE.ink))
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1,
                               wrap_text=True)
    for offset in range(span):
        target = ws.cell(row, col + offset)
        target.fill = fill(PALETTE.sand_deep)
        target.border = Border(bottom=side(PALETTE.paper_warm),
                               top=side(PALETTE.paper_warm))


def input_cell(ws, row: int, col: int, *, span: int = 1, number_format: str = "General",
               value=None, centre: bool = False) -> None:
    """A cell the user is expected to fill in: pale gold, gold underline, unlocked."""
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row,
                       end_column=col + span - 1)
    for offset in range(span):
        cell = ws.cell(row, col + offset)
        cell.fill = fill(PALETTE.field)
        cell.border = FIELD_BORDER
        cell.font = Font(name=FONT_BODY, size=10, color=argb(PALETTE.ink))
        cell.alignment = WRAP_CENTRE if centre else Alignment(
            horizontal="left", vertical="center", indent=1, wrap_text=True)
        cell.number_format = number_format
        cell.protection = Protection(locked=False)
    if value is not None:
        ws.cell(row, col).value = value


def field_row(ws, row: int, *, label_col: int, label_text: str, label_span: int,
              value_col: int, value_span: int, number_format: str = "General",
              value=None, height: float = 18) -> None:
    label(ws, row, label_col, label_text, span=label_span)
    input_cell(ws, row, value_col, span=value_span, number_format=number_format,
               value=value)
    ws.row_dimensions[row].height = height


def formula_cell(ws, row: int, col: int, formula: str, *, span: int = 1,
                 number_format: str = "General", bold: bool = False,
                 colour: str = PALETTE.ink, background: str = PALETTE.sand) -> None:
    """A derived cell — visually distinct from an input so nobody types over it."""
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row,
                       end_column=col + span - 1)
    for offset in range(span):
        cell = ws.cell(row, col + offset)
        cell.fill = fill(background)
        cell.border = HAIRLINE
        cell.font = Font(name=FONT_BODY, size=10, bold=bold, color=argb(colour))
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1,
                                   wrap_text=True)
        cell.number_format = number_format
    ws.cell(row, col).value = formula


def table_header(ws, row: int, headers: list[str], brand: BrandProfile,
                 start_col: int = 1) -> None:
    for index, text in enumerate(headers):
        cell = ws.cell(row, start_col + index)
        cell.value = text
        cell.font = Font(name=FONT_BODY, size=8.5, bold=True, color=argb(PALETTE.paper))
        cell.fill = fill(brand.primary)
        cell.alignment = WRAP_CENTRE
        cell.border = Border(bottom=side(brand.accent, "medium"))
    ws.row_dimensions[row].height = 28


def table_body(ws, first_row: int, last_row: int, columns: int,
               formats: dict[int, str] | None = None, start_col: int = 1,
               zebra: bool = True) -> None:
    formats = formats or {}
    for row in range(first_row, last_row + 1):
        tint = PALETTE.paper_warm if (row - first_row) % 2 == 0 or not zebra else PALETTE.sand
        ws.row_dimensions[row].height = 17
        for index in range(columns):
            cell = ws.cell(row, start_col + index)
            cell.fill = fill(tint)
            cell.border = HAIRLINE
            cell.font = Font(name=FONT_BODY, size=10, color=argb(PALETTE.ink))
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.number_format = formats.get(index, "General")
            cell.protection = Protection(locked=False)


def total_row(ws, row: int, columns: int, brand: BrandProfile, *, label_text: str,
              label_span: int, formulas: dict[int, str],
              formats: dict[int, str] | None = None, start_col: int = 1) -> None:
    formats = formats or {}
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row,
                   end_column=start_col + label_span - 1)
    for index in range(columns):
        cell = ws.cell(row, start_col + index)
        cell.fill = fill(PALETTE.gold_tint)
        cell.border = Border(top=side(brand.accent, "medium"),
                             bottom=side(brand.accent, "medium"))
        cell.font = Font(name=FONT_BODY, size=10, bold=True, color=argb(PALETTE.gold_dark))
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.number_format = formats.get(index, "General")
    ws.cell(row, start_col).value = label_text
    for index, formula in formulas.items():
        ws.cell(row, start_col + index).value = formula
    ws.row_dimensions[row].height = 20


def note(ws, cell_range: str, title: str, body: str, *, tone: str = "neutral") -> None:
    """Guidance panel — the spreadsheet twin of ``docx_kit.note_card``."""
    tones = {
        "neutral": (PALETTE.sand, PALETTE.line, PALETTE.ink_soft),
        "brand": (PALETTE.gold_pale, PALETTE.gold_mid, PALETTE.gold_dark),
        "info": (PALETTE.azure_soft, "BFE3F7", PALETTE.azure),
        "alert": (PALETTE.alert_soft, "E8C4B4", PALETTE.alert),
    }
    background, border, accent = tones.get(tone, tones["neutral"])
    ws.merge_cells(cell_range)
    first = cell_range.split(":")[0]
    cell = ws[first]
    cell.value = f"{title.upper()}\n{body}"
    cell.fill = fill(background)
    cell.font = Font(name=FONT_BODY, size=9, color=argb(PALETTE.ink))
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True,
                               indent=1)
    start_row = int("".join(c for c in first if c.isdigit()))
    end_row = int("".join(c for c in cell_range.split(":")[1] if c.isdigit()))
    last = _last_col(cell_range)
    for row in range(start_row, end_row + 1):
        for col in range(1, last + 1):
            target = ws.cell(row, col)
            target.fill = fill(background)
            edges = {}
            if row == start_row:
                edges["top"] = side(border)
            if row == end_row:
                edges["bottom"] = side(border)
            if col == 1:
                edges["left"] = side(accent, "thick")
            if col == last:
                edges["right"] = side(border)
            target.border = Border(**edges)


def kpi_tile(ws, row: int, col: int, span: int, title: str, formula: str,
             brand: BrandProfile, *, number_format: str = MONEY) -> None:
    """Two-row stat tile: obsidian caption over a large gold figure."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    head = ws.cell(row, col)
    head.value = title
    head.font = Font(name=FONT_BODY, size=8, bold=True, color=argb(brand.accent))
    head.fill = fill(brand.primary)
    head.alignment = Alignment(horizontal="center", vertical="center")
    for offset in range(span):
        ws.cell(row, col + offset).fill = fill(brand.primary)
    ws.row_dimensions[row].height = 18

    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2,
                   end_column=col + span - 1)
    value = ws.cell(row + 1, col)
    value.value = formula
    value.font = Font(name=FONT_DISPLAY, size=15, bold=True, color=argb(PALETTE.ink))
    value.fill = fill(PALETTE.gold_pale)
    value.alignment = Alignment(horizontal="center", vertical="center")
    value.number_format = number_format
    for extra_row in (row + 1, row + 2):
        ws.row_dimensions[extra_row].height = 15
        for offset in range(span):
            cell = ws.cell(extra_row, col + offset)
            cell.fill = fill(PALETTE.gold_pale)
            edges = {"left": side(brand.accent), "right": side(brand.accent)}
            if extra_row == row + 2:
                edges["bottom"] = side(brand.accent)
            cell.border = Border(**edges)


def dropdown(ws, source_range: str, target: str, *, prompt: str = "") -> None:
    validation = DataValidation(type="list", formula1=source_range, allow_blank=True,
                                showDropDown=False)
    validation.error = "Choose a value from the list."
    validation.errorTitle = "Invalid entry"
    if prompt:
        validation.prompt = prompt
        validation.promptTitle = "Select"
    ws.add_data_validation(validation)
    validation.add(target)


def non_negative(ws, target: str, *, message: str = "Enter a number of zero or more.") -> None:
    validation = DataValidation(type="decimal", operator="greaterThanOrEqual",
                                formula1="0", allow_blank=True)
    validation.error = message
    validation.errorTitle = "Invalid amount"
    ws.add_data_validation(validation)
    validation.add(target)


def flag_missing(ws, target: str) -> None:
    """Amber tint on a required input that is still blank."""
    first = target.split(":")[0]
    ws.conditional_formatting.add(
        target,
        FormulaRule(formula=[f'ISBLANK({first})'], fill=fill("FDF3D6"), stopIfTrue=False),
    )


def flag_negative(ws, target: str) -> None:
    ws.conditional_formatting.add(
        target,
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(name=FONT_BODY, bold=True, color=argb(PALETTE.alert))),
    )


def print_setup(ws, brand: BrandProfile, *, title_rows: str | None = None,
                landscape: bool = False, area: str | None = None,
                fit_width: int = 1) -> None:
    """A4, fit-to-width, repeated headers and a branded print footer."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = 0.55
    ws.page_margins.bottom = 0.55
    ws.page_margins.header = ws.page_margins.footer = 0.25
    ws.print_options.horizontalCentered = True
    if title_rows:
        ws.print_title_rows = title_rows
    if area:
        ws.print_area = area
    ws.oddHeader.left.text = "&\"Calibri\"&8&K7C5E13" + brand.confidentiality
    ws.oddHeader.right.text = "&\"Calibri\"&8&K9A8D7C&A"
    ws.oddFooter.left.text = f"&\"Calibri\"&8&K9A8D7C{brand.platform_note}  |  v{brand.version}"
    ws.oddFooter.right.text = "&\"Calibri\"&8&K9A8D7CPage &P of &N"


def widths(ws, mapping: dict[str, float]) -> None:
    for column, width in mapping.items():
        ws.column_dimensions[column].width = width


def freeze(ws, cell: str) -> None:
    ws.freeze_panes = cell


def blank_row(ws, row: int, height: float = 6) -> None:
    ws.row_dimensions[row].height = height
