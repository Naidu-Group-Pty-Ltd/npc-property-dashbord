#!/usr/bin/env python3
"""Structural checks on the generated Finance Portal templates.

Guards the things that are easy to break silently when the builder changes:
the workbook must still open and its cross-sheet formulas must still point at
the rows they describe.

**The two referral agreements are no longer generated.** They are shipped as
supplied by their author and verified clause by clause in TypeScript by
``src/lib/agreements/__tests__/agreementTemplateFiles.spec.ts``, which is the
authority on them. What is checked here instead is that nothing has started
writing a SECOND copy of either agreement into this directory — three
typesettings of the same instrument is what that suite exists to prevent.

    python3 scripts/finance-portal-templates/verify_templates.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from docx import Document
from docx.oxml.ns import qn
from openpyxl import load_workbook

OUT = Path(__file__).resolve().parents[2] / "public" / "templates" / "finance-portal"

#: The documents this directory is allowed to hold, and who owns each. A Word
#: file here that is not on this list means a builder has started generating an
#: agreement again, beside the one people actually download.
SHIPPED_AGREEMENTS = {
    "Strategic_Property_Referral_Agreement.docx",
    "Finance_Referral_and_Commission_Agreement.docx",
}

XLSX_FILE = "Aurixa_White_Label_Client_Fact_Find.xlsx"
EXPECTED_SHEETS = ["Start Here", "White Label Setup", "Client Fact Find",
                   "Living Expenses", "Client Form Output", "Lists"]

#: The summary sheet must read the fact find at these rows. Each entry is
#: (summary cell, label the summary prints, fact-find row it must reference).
OUTPUT_BINDINGS = [
    ("C27", "Employer / Business", 28),
    ("C28", "Role / Position", 29),
    ("C29", "Employment Type", 27),
    ("C30", "Start Date", 31),
]

failures: list[str] = []


def check(condition: bool, message: str) -> None:
    if not condition:
        failures.append(message)


def all_text(document: Document) -> str:
    return "\n".join(node.text or "" for node in document.element.body.iter(qn("w:t")))


def verify_docx() -> None:
    """No agreement may be generated into this directory."""
    present = {path.name for path in OUT.glob("*.docx")}
    for filename in sorted(SHIPPED_AGREEMENTS):
        path = OUT / filename
        check(path.exists(), f"{filename}: missing — the shipped agreement is not here")
        if not path.exists():
            continue
        text = all_text(Document(path))
        # Merge tokens must survive, or the pack stops being white-label.
        check("<<COMPANY NAME>>" in text, f"{filename}: <<COMPANY NAME>> token missing")
        check("<<INSERT>>" in text, f"{filename}: <<INSERT>> token missing")

    extra = sorted(present - SHIPPED_AGREEMENTS)
    check(not extra,
          "a builder has written Word document(s) beside the shipped agreements: "
          f"{extra}. Nothing may generate an agreement into this directory — see "
          "_shared/agreements/templateFiles.pure.ts")
    print(f"  ✓ {len(SHIPPED_AGREEMENTS)} shipped agreement(s) present, no generated copies")


def verify_xlsx() -> None:
    path = OUT / XLSX_FILE
    check(path.exists(), f"{XLSX_FILE}: missing")
    if not path.exists():
        return
    wb = load_workbook(path)
    check(wb.sheetnames == EXPECTED_SHEETS,
          f"{XLSX_FILE}: sheets are {wb.sheetnames}, expected {EXPECTED_SHEETS}")

    fact = wb["Client Fact Find"]
    output = wb["Client Form Output"]
    expenses = wb["Living Expenses"]

    # Employment labels on the fact find must sit where the summary expects.
    for cell, label, source_row in OUTPUT_BINDINGS:
        actual_label = fact[f"A{source_row}"].value
        check(actual_label == label,
              f"{XLSX_FILE}: fact find A{source_row} is '{actual_label}', expected '{label}'")
        formula = output[cell].value or ""
        check(f"C{source_row}" in str(formula),
              f"{XLSX_FILE}: summary {cell} does not read fact find row {source_row} "
              f"(got {formula!r})")

    # Income total must span exactly the five income rows.
    check("SUM('Client Fact Find'!C32:C36)" in str(output["C31"].value),
          f"{XLSX_FILE}: income total does not sum rows 32–36 (got {output['C31'].value!r})")
    for row, label in ((32, "Base Salary (Annual)"), (36, "Other Taxable Income")):
        check(fact[f"A{row}"].value == label,
              f"{XLSX_FILE}: fact find A{row} is {fact[f'A{row}'].value!r}, expected {label!r}")

    # Living-situation dropdown must be on the living-situation row, not e-mail.
    living_targets = [
        str(dv.sqref) for dv in fact.data_validations.dataValidation
        if dv.formula1 and "$E$2:$E$" in str(dv.formula1)
    ]
    joined = " ".join(living_targets)
    check("C20" in joined and "H20" in joined,
          f"{XLSX_FILE}: living-situation dropdown is on {joined or 'no row'}, expected row 20")
    check("C16" not in joined,
          f"{XLSX_FILE}: living-situation dropdown is still attached to the e-mail row")

    # The summary cover must show the tagline, never the brand hex.
    check("C8" in str(output["A4"].value),
          f"{XLSX_FILE}: summary tagline reads {output['A4'].value!r}, expected setup C8")

    # Living expenses must ship at zero across the board.
    stray = [
        f"C{row}={expenses[f'C{row}'].value}"
        for row in range(5, 55)
        if expenses[f"C{row}"].value not in (0, None)
    ]
    check(not stray, f"{XLSX_FILE}: living expenses seeded with non-zero values: {stray}")

    check("Orixa" not in str(wb["White Label Setup"]["F12"].value or ""),
          f"{XLSX_FILE}: 'Orixa' typo still present")

    for sheet in ("Client Fact Find", "Living Expenses", "Client Form Output"):
        check(wb[sheet].page_setup.fitToWidth == 1,
              f"{XLSX_FILE}: {sheet} is not set to fit one page wide")
    check(wb["Lists"].sheet_state == "hidden", f"{XLSX_FILE}: Lists sheet should be hidden")

    print(f"  ✓ {XLSX_FILE} — {len(wb.sheetnames)} sheets, bindings verified")


def main() -> int:
    print("Verifying Finance Portal templates…")
    verify_docx()
    verify_xlsx()
    if failures:
        print(f"\n{len(failures)} problem(s):")
        for failure in failures:
            print(f"  ✗ {failure}")
        return 1
    print("\nAll checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
