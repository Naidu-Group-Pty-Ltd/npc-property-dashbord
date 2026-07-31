#!/usr/bin/env python3
"""Build the White-Label Client Fact Find workbook.

Rebuilds the source workbook with the Aurixa design system, keeping every
original sheet, field and formula intent, and correcting the defects carried
over from the first draft:

* ``Client Form Output`` read the employment block three rows too low, so
  "Employer" showed the start date, "Role" showed the base salary and the
  income total summed two income rows plus the assets header.
* The living-situation dropdown was attached to the e-mail row (16) rather
  than the living-situation row (20).
* The output cover printed the primary-colour hex string where the tagline
  belonged.
* ``Living Expenses`` shipped with a stray 600 in Registration; every other
  line was zero.
* The setup sheet referred to the platform as "Orixa".

Run via ``build_all.py``.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Protection, Side
from openpyxl.utils import get_column_letter

from aurixa_brand import DEFAULT_BRAND, PALETTE, BrandProfile
from xlsx_kit import (
    DATE,
    FONT_BODY,
    FONT_DISPLAY,
    MONEY,
    PERCENT,
    argb,
    banner,
    blank_row,
    dropdown,
    field_row,
    fill,
    flag_missing,
    flag_negative,
    formula_cell,
    freeze,
    input_cell,
    kpi_tile,
    label,
    non_negative,
    note,
    print_setup,
    section_bar,
    subhead,
    table_body,
    table_header,
    total_row,
    widths,
)

SHEET_START = "Start Here"
SHEET_SETUP = "White Label Setup"
SHEET_FACT = "Client Fact Find"
SHEET_EXPENSES = "Living Expenses"
SHEET_OUTPUT = "Client Form Output"
SHEET_LISTS = "Lists"

# Reference lists driving every dropdown. Kept on a hidden sheet so the visible
# tabs stay clean but the values remain editable per organisation.
LISTS: dict[str, list[str]] = {
    "Title": ["Mr", "Mrs", "Ms", "Miss", "Dr", "Other"],
    "Gender": ["Female", "Male", "Non-binary", "Prefer not to say", "Other"],
    "MaritalStatus": ["Single", "Married", "De facto", "Separated", "Divorced", "Widowed"],
    "Residency": ["Australian citizen", "Permanent resident", "Temporary visa",
                  "New Zealand citizen", "Non-resident", "Other"],
    "LivingSituation": ["Owned", "Mortgaged", "Renting", "Boarding",
                        "Living with parents", "Other"],
    "EmploymentType": ["PAYG full time", "PAYG part time", "PAYG casual", "Contract",
                       "Self-employed", "Company director", "Retired", "Home duties",
                       "Student", "Unemployed"],
    "Owner": ["Applicant 1", "Applicant 2", "Joint", "Other"],
    "AssetType": ["Owner-occupied property", "Investment property", "Vehicle",
                  "Savings / cash", "Term deposit", "Shares / managed funds",
                  "Superannuation", "Business interest", "Contents", "Other"],
    "LiabilityType": ["Home loan", "Investment loan", "Personal loan", "Car loan",
                      "Credit card", "Store card", "Overdraft", "Buy now pay later",
                      "HECS / HELP", "Tax debt", "Other"],
    "YesNo": ["Yes", "No"],
    "Frequency": ["Weekly", "Fortnightly", "Monthly", "Quarterly", "Annually"],
}

EXPENSE_ITEMS: list[tuple[str, str]] = [
    ("Childcare & Support", "Childcare"),
    ("Childcare & Support", "Child Maintenance"),
    ("Education", "Public School Costs"),
    ("Education", "Private School Costs"),
    ("Education", "Higher Education / Vocational Training"),
    ("Groceries", "Groceries"),
    ("Primary Residence", "Electricity & Gas"),
    ("Primary Residence", "Council Rates"),
    ("Primary Residence", "Water & Sewer"),
    ("Primary Residence", "Body Corporate"),
    ("Primary Residence", "Home Repairs"),
    ("Primary Residence", "Furnishings & Electrical"),
    ("Insurance", "Building Insurance"),
    ("Insurance", "Contents Insurance"),
    ("Insurance", "Health Insurance"),
    ("Insurance", "Income Protection"),
    ("Insurance", "Life Insurance"),
    ("Insurance", "Vehicle Insurance"),
    ("Investment Property", "Rates, Utilities & Body Corporate"),
    ("Investment Property", "Repairs & Maintenance"),
    ("Investment Property", "Insurance"),
    ("Secondary Residence", "Rates, Utilities & Body Corporate"),
    ("Secondary Residence", "Repairs & Maintenance"),
    ("Secondary Residence", "Insurance"),
    ("Medical", "Medical & Health"),
    ("Medical", "Natural Therapies"),
    ("Housing", "Rent"),
    ("Housing", "Board"),
    ("Personal Care", "Clothing & Footwear"),
    ("Personal Care", "Cosmetics / Haircare"),
    ("Personal Care", "Dry Cleaning"),
    ("Recreation", "Pets"),
    ("Recreation", "Alcohol / Tobacco"),
    ("Recreation", "Cinema / Concerts / Memberships"),
    ("Recreation", "Dining Out"),
    ("Recreation", "Gym / Sports"),
    ("Recreation", "Travel & Holidays"),
    ("Recreation", "Gifts & Miscellaneous"),
    ("Recreation", "Gambling"),
    ("Communications", "Home / Mobile Phone"),
    ("Communications", "Internet / Pay TV / Streaming"),
    ("Transport", "Petrol"),
    ("Transport", "Registration"),
    ("Transport", "Vehicle Maintenance"),
    ("Transport", "Public Transport"),
    ("Transport", "Taxi / Ride Sharing"),
    ("Transport", "Tolls / Parking"),
    ("Other", "Regular Donations"),
    ("Other", "Voluntary Superannuation"),
    ("Other", "Other Regular Expense"),
]

# Row anchors on the Client Fact Find sheet. Named so the summary sheet can
# never drift out of sync with the form again.
PERSONAL_FIELDS = [
    ("Title", "list", "Title"),
    ("First Name", "text", None),
    ("Middle Name", "text", None),
    ("Surname", "text", None),
    ("Date of Birth", "date", None),
    ("Gender", "list", "Gender"),
    ("Marital Status", "list", "MaritalStatus"),
    ("Residency Status", "list", "Residency"),
    ("Number of Dependants", "int", None),
    ("Mobile", "text", None),
    ("Email", "text", None),
]
ADDRESS_FIELDS = [
    ("Current Address", "text", None),
    ("Living Situation", "list", "LivingSituation"),
    ("Date Moved In", "date", None),
    ("Previous Address", "text", None),
    ("Previous Living Situation", "list", "LivingSituation"),
    ("Previous Date Moved In", "date", None),
]
EMPLOYMENT_FIELDS = [
    ("Employment Type", "list", "EmploymentType"),
    ("Employer / Business", "text", None),
    ("Role / Position", "text", None),
    ("Employer Address", "text", None),
    ("Start Date", "date", None),
    ("Base Salary (Annual)", "money", None),
    ("Bonus", "money", None),
    ("Commission", "money", None),
    ("Overtime", "money", None),
    ("Other Taxable Income", "money", None),
]

ROW_PERSONAL = 5          # section bar; fields run 6..16
ROW_ADDRESS = 18          # fields 19..24
ROW_EMPLOYMENT = 26       # fields 27..36
ROW_ASSETS = 38           # header 39, data 40..49, total 50
ROW_LIABILITIES = 51      # header 52, data 53..60, total 61
ASSET_FIRST, ASSET_LAST, ASSET_TOTAL = 40, 49, 50
LIAB_FIRST, LIAB_LAST, LIAB_TOTAL = 53, 60, 61
INCOME_FIRST, INCOME_LAST = 32, 36
EXPENSE_FIRST = 5
EXPENSE_LAST = EXPENSE_FIRST + len(EXPENSE_ITEMS) - 1
EXPENSE_TOTAL = EXPENSE_LAST + 1


def _blank_safe(reference: str) -> str:
    """Pass a cell through, but print nothing when the source is empty.

    Without this an unfilled text field renders as ``0`` and an unfilled date
    renders as 30/12/1899 on the printed summary.
    """
    return f'IF({reference}="","",{reference})'


def _list_ref(name: str) -> str:
    """Absolute reference to a named list column on the hidden Lists sheet."""
    index = list(LISTS).index(name) + 1
    column = get_column_letter(index)
    return f"'{SHEET_LISTS}'!${column}$2:${column}${len(LISTS[name]) + 1}"


# ==========================================================================
# Sheets
# ==========================================================================

def build_lists(wb: Workbook, brand: BrandProfile) -> None:
    ws = wb.create_sheet(SHEET_LISTS)
    for index, (name, values) in enumerate(LISTS.items(), start=1):
        header = ws.cell(1, index)
        header.value = name
        header.font = Font(name=FONT_BODY, size=10, bold=True, color=argb(PALETTE.paper))
        header.fill = fill(brand.primary)
        ws.column_dimensions[get_column_letter(index)].width = 24
        for offset, value in enumerate(values, start=2):
            ws.cell(offset, index).value = value
    ws.sheet_state = "hidden"


def build_start_here(wb: Workbook, brand: BrandProfile) -> None:
    ws = wb.create_sheet(SHEET_START, 0)
    widths(ws, {"A": 3, "B": 22, "C": 34, "D": 26, "E": 10, "F": 42})
    banner(ws, "A1:F3", "CLIENT FINANCIAL POSITION & FACT FIND", brand, size=17)

    ws["A5"] = "WHITE-LABEL WORKBOOK  ·  START HERE"
    ws["A5"].font = Font(name=FONT_BODY, size=9, bold=True, color=argb(PALETTE.gold_dark))
    blank_row(ws, 4, 8)

    note(ws, "A6:F8", "How this workbook fits together",
         "Complete the tabs left to right. 'White Label Setup' carries your organisation's "
         "name, contact details and brand colours; every other tab reads from it. "
         "'Client Fact Find' and 'Living Expenses' are the data-entry forms. "
         "'Client Form Output' is the print-ready summary — it is entirely formula-driven, "
         "so never type into it.",
         tone="brand")

    blank_row(ws, 9, 8)
    section_bar(ws, 10, 6, "TAB GUIDE", brand, kicker="What each sheet is for")
    table_header(ws, 11, ["Tab", "Purpose", "Who completes it", "Print?", "Notes"], brand,
                 start_col=2)
    guide = [
        (SHEET_SETUP, "Organisation branding and document settings",
         "Partner administrator", "No", "Complete once per organisation."),
        (SHEET_FACT, "Applicant details, employment, assets and liabilities",
         "Client, with adviser support", "Yes", "Yellow cells are inputs."),
        (SHEET_EXPENSES, "Monthly household expenditure by category",
         "Client", "Yes", "Enter monthly figures; annual is calculated."),
        (SHEET_OUTPUT, "Print-ready summary and client declaration",
         "Generated automatically", "Yes", "Formula-driven — do not overtype."),
    ]
    table_body(ws, 12, 15, 5, start_col=2)
    for offset, row_values in enumerate(guide):
        ws.row_dimensions[12 + offset].height = 30
        for index, value in enumerate(row_values):
            cell = ws.cell(12 + offset, 2 + index)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1, wrap_text=True)

    blank_row(ws, 16, 8)
    section_bar(ws, 17, 6, "COMPLETION CHECKLIST", brand, kicker="Before you issue or export")
    checklist = [
        "Set the organisation name, contact details and brand colours on 'White Label Setup'.",
        "Replace the logo placeholder with the partner mark.",
        "Confirm the confidentiality label and disclaimer wording.",
        "Complete every applicant, employment, asset and liability row that applies.",
        "Enter monthly living expenses — leave inapplicable lines at zero.",
        "Review 'Client Form Output' for blank or negative figures.",
        "Have the client sign the declaration, then export to PDF.",
    ]
    table_body(ws, 18, 17 + len(checklist), 5, start_col=2)
    for offset, item in enumerate(checklist):
        cell = ws.cell(18 + offset, 2)
        cell.value = f"☐   {item}"
        ws.merge_cells(start_row=18 + offset, start_column=2,
                       end_row=18 + offset, end_column=6)

    last = 18 + len(checklist)
    blank_row(ws, last + 1, 8)
    note(ws, f"A{last + 2}:F{last + 4}", "Data handling",
         "This workbook holds personal and financial information. Store it in a "
         "restricted-access location, share it only through an approved secure channel, "
         "and retain it for the period required by your privacy policy and compliance "
         "framework.",
         tone="alert")

    print_setup(ws, brand, area=f"A1:F{last + 4}")
    ws.sheet_view.showGridLines = False


def build_setup(wb: Workbook, brand: BrandProfile) -> None:
    ws = wb.create_sheet(SHEET_SETUP)
    widths(ws, {"A": 3, "B": 26, "C": 34, "D": 4, "E": 26, "F": 34})
    banner(ws, "A1:F3", "WHITE-LABEL CONFIGURATION", brand, size=17)
    blank_row(ws, 4, 8)

    section_bar(ws, 5, 6, "ORGANISATION IDENTITY", brand,
                kicker="Shown on every tab and in the exported PDF")

    identity = [
        ("Organisation Name", "NPC Services"),
        ("Trading Name / Division", ""),
        ("Tagline", "YOUR TRUSTED PROPERTY & FINANCE PARTNER"),
        ("Website", "www.yourorganisation.com.au"),
        ("Email", "info@yourorganisation.com.au"),
        ("Phone", "(00) 0000 0000"),
        ("Business Address", "Insert business address"),
        ("ABN / ACN", ""),
    ]
    document = [
        ("Document Title", "Client Financial Position & Fact Find"),
        ("Confidentiality Label", "CONFIDENTIAL"),
        ("Prepared By", ""),
        ("Version", "1.0"),
        ("Primary Colour (Hex)", f"#{brand.primary}"),
        ("Accent Colour (Hex)", f"#{brand.accent}"),
        ("Logo Placeholder", "Replace this area with your logo"),
        ("Disclaimer", "Insert licensing or credit-guide wording, or leave blank"),
    ]

    for offset, ((left_label, left_value), (right_label, right_value)) in enumerate(
            zip(identity, document)):
        row = 6 + offset
        field_row(ws, row, label_col=2, label_text=left_label, label_span=1,
                  value_col=3, value_span=1, value=left_value or None, height=19)
        label(ws, row, 5, right_label)
        input_cell(ws, row, 6, value=right_value or None)

    flag_missing(ws, "C6")
    flag_missing(ws, "F6")

    blank_row(ws, 14, 8)
    note(ws, "A15:F17", "Colour fields",
         "The hex values above are recorded for reference and are applied automatically "
         "when the Aurixa Finance Portal generates a branded export. To re-skin the "
         "workbook by hand, select a banner row and apply the same colour from Excel's "
         "fill picker — the semantic colours (green, amber, red) should stay as they are.",
         tone="info")

    blank_row(ws, 18, 8)
    note(ws, "A19:F21", "Implementation note",
         "This workbook is intentionally organisation-neutral. The Aurixa Systems Client "
         "Forms module can map each editable field to its database, apply the selected "
         "organisation's brand settings, and generate the downloadable 'Client Form "
         "Output' without displaying any external finance-partner branding.",
         tone="brand")

    print_setup(ws, brand, area="A1:F21")
    ws.sheet_view.showGridLines = False


def _applicant_block(ws, brand: BrandProfile, start_row: int, fields, column_base: int,
                     validations: list[tuple[str, str]]) -> None:
    """Render one applicant's field stack. ``column_base`` is 1 (A) or 6 (F)."""
    formats = {"date": DATE, "money": MONEY, "int": "0"}
    for offset, (name, kind, list_name) in enumerate(fields):
        row = start_row + offset
        field_row(ws, row, label_col=column_base, label_text=name, label_span=2,
                  value_col=column_base + 2, value_span=3,
                  number_format=formats.get(kind, "General"))
        if kind == "list" and list_name:
            target = (f"{get_column_letter(column_base + 2)}{row}:"
                      f"{get_column_letter(column_base + 4)}{row}")
            validations.append((list_name, target))
        if kind == "money":
            non_negative(ws, f"{get_column_letter(column_base + 2)}{row}")


def build_fact_find(wb: Workbook, brand: BrandProfile) -> None:
    ws = wb.create_sheet(SHEET_FACT)
    widths(ws, {"A": 15, "B": 14, "C": 13, "D": 12, "E": 12,
                "F": 15, "G": 14, "H": 13, "I": 12, "J": 12})
    banner(ws, "A1:J2", "CLIENT FACT FIND", brand, size=18)

    brandline = ws.cell(3, 1)
    ws.merge_cells("A3:J3")
    brandline.value = (f"='{SHEET_SETUP}'!C6&\" | \"&'{SHEET_SETUP}'!F6"
                       f"&\" | \"&'{SHEET_SETUP}'!F7")
    brandline.font = Font(name=FONT_BODY, size=9, bold=True, color=argb(PALETTE.obsidian))
    brandline.fill = fill(brand.accent)
    brandline.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 18
    for col in range(1, 11):
        ws.cell(3, col).fill = fill(brand.accent)
    blank_row(ws, 4, 8)

    validations: list[tuple[str, str]] = []

    section_bar(ws, ROW_PERSONAL, 10, "PERSONAL DETAILS", brand,
                kicker="Applicant 1 (columns A–E)   ·   Applicant 2 (columns F–J)")
    _applicant_block(ws, brand, ROW_PERSONAL + 1, PERSONAL_FIELDS, 1, validations)
    _applicant_block(ws, brand, ROW_PERSONAL + 1, PERSONAL_FIELDS, 6, validations)
    for coordinate in ("C7", "C9", "H7", "H9"):
        flag_missing(ws, coordinate)
    blank_row(ws, 17, 8)

    section_bar(ws, ROW_ADDRESS, 10, "ADDRESS HISTORY", brand,
                kicker="Include the previous address where the current one is under three years")
    _applicant_block(ws, brand, ROW_ADDRESS + 1, ADDRESS_FIELDS, 1, validations)
    _applicant_block(ws, brand, ROW_ADDRESS + 1, ADDRESS_FIELDS, 6, validations)
    blank_row(ws, 25, 8)

    section_bar(ws, ROW_EMPLOYMENT, 10, "CURRENT EMPLOYMENT & INCOME", brand,
                kicker="Gross annual amounts")
    _applicant_block(ws, brand, ROW_EMPLOYMENT + 1, EMPLOYMENT_FIELDS, 1, validations)
    _applicant_block(ws, brand, ROW_EMPLOYMENT + 1, EMPLOYMENT_FIELDS, 6, validations)
    blank_row(ws, 37, 8)

    for list_name, target in validations:
        dropdown(ws, _list_ref(list_name), target)

    # ---------------------------------------------------------------- assets
    section_bar(ws, ROW_ASSETS, 10, "ASSETS", brand,
                kicker="Include any loan secured against each asset")
    table_header(ws, ROW_ASSETS + 1, [
        "Asset Type", "Description / Address", "Owner", "Current Value",
        "Rental / Other Income", "Financial Institution", "Loan Balance",
        "Monthly Repayment", "Interest Rate", "Maturity Date",
    ], brand)
    table_body(ws, ASSET_FIRST, ASSET_LAST, 10, formats={
        3: MONEY, 4: MONEY, 6: MONEY, 7: MONEY, 8: PERCENT, 9: DATE,
    })
    dropdown(ws, _list_ref("AssetType"), f"A{ASSET_FIRST}:A{ASSET_LAST}")
    dropdown(ws, _list_ref("Owner"), f"C{ASSET_FIRST}:C{ASSET_LAST}")
    total_row(
        ws, ASSET_TOTAL, 10, brand,
        label_text="TOTAL ASSETS", label_span=3,
        formulas={
            3: f"=SUM(D{ASSET_FIRST}:D{ASSET_LAST})",
            4: f"=SUM(E{ASSET_FIRST}:E{ASSET_LAST})",
            6: f"=SUM(G{ASSET_FIRST}:G{ASSET_LAST})",
            7: f"=SUM(H{ASSET_FIRST}:H{ASSET_LAST})",
        },
        formats={3: MONEY, 4: MONEY, 6: MONEY, 7: MONEY},
    )

    # ----------------------------------------------------------- liabilities
    section_bar(ws, ROW_LIABILITIES, 10, "OTHER LIABILITIES", brand,
                kicker="Debts not already secured against an asset above")
    table_header(ws, ROW_LIABILITIES + 1, [
        "Liability Type", "Lender", "Account / Description", "Owner",
        "Limit / Original Amount", "Current Balance", "Monthly Repayment",
        "Interest Rate", "Remaining Term", "Notes",
    ], brand)
    table_body(ws, LIAB_FIRST, LIAB_LAST, 10, formats={
        4: MONEY, 5: MONEY, 6: MONEY, 7: PERCENT,
    })
    dropdown(ws, _list_ref("LiabilityType"), f"A{LIAB_FIRST}:A{LIAB_LAST}")
    dropdown(ws, _list_ref("Owner"), f"D{LIAB_FIRST}:D{LIAB_LAST}")
    total_row(
        ws, LIAB_TOTAL, 10, brand,
        label_text="TOTAL OTHER LIABILITIES", label_span=4,
        formulas={
            4: f"=SUM(E{LIAB_FIRST}:E{LIAB_LAST})",
            5: f"=SUM(F{LIAB_FIRST}:F{LIAB_LAST})",
            6: f"=SUM(G{LIAB_FIRST}:G{LIAB_LAST})",
        },
        formats={4: MONEY, 5: MONEY, 6: MONEY},
    )

    blank_row(ws, LIAB_TOTAL + 1, 8)
    note(ws, f"A{LIAB_TOTAL + 2}:J{LIAB_TOTAL + 4}", "Completing this form",
         "Yellow cells are for you to complete; grey totals calculate themselves. "
         "Leave a row blank rather than entering zero if it does not apply. Where a loan "
         "is secured against an asset, record it on the asset row so it is not counted "
         "twice.",
         tone="info")

    freeze(ws, "A5")
    print_setup(ws, brand, title_rows="1:3", landscape=True,
                area=f"A1:J{LIAB_TOTAL + 4}")
    ws.sheet_view.showGridLines = False


def build_expenses(wb: Workbook, brand: BrandProfile) -> None:
    ws = wb.create_sheet(SHEET_EXPENSES)
    widths(ws, {"A": 24, "B": 40, "C": 16, "D": 16, "E": 26,
                "F": 3, "G": 24, "H": 16})
    banner(ws, "A1:E2", "MONTHLY LIVING EXPENSES", brand, size=18)
    blank_row(ws, 3, 8)

    table_header(ws, 4, ["Category", "Expense Item", "Monthly Amount",
                         "Annual (auto)", "Notes"], brand)
    table_body(ws, EXPENSE_FIRST, EXPENSE_LAST, 5,
               formats={2: MONEY, 3: MONEY})

    for offset, (category, item) in enumerate(EXPENSE_ITEMS):
        row = EXPENSE_FIRST + offset
        ws.cell(row, 1).value = category
        ws.cell(row, 2).value = item
        ws.cell(row, 3).value = 0
        annual = ws.cell(row, 4)
        annual.value = f"=C{row}*12"
        annual.fill = fill(PALETTE.sand)
        annual.font = Font(name=FONT_BODY, size=10, color=argb(PALETTE.ink_soft))
        for column in (1, 2, 4):
            ws.cell(row, column).protection = Protection(locked=True)
        for column in (3, 5):
            cell = ws.cell(row, column)
            cell.fill = fill(PALETTE.field)
            cell.protection = Protection(locked=False)

    non_negative(ws, f"C{EXPENSE_FIRST}:C{EXPENSE_LAST}")
    total_row(
        ws, EXPENSE_TOTAL, 5, brand,
        label_text="TOTAL MONTHLY LIVING EXPENSES", label_span=2,
        formulas={
            2: f"=SUM(C{EXPENSE_FIRST}:C{EXPENSE_LAST})",
            3: f"=SUM(D{EXPENSE_FIRST}:D{EXPENSE_LAST})",
        },
        formats={2: MONEY, 3: MONEY},
    )

    # Category roll-up beside the detail, so the client can sanity-check totals.
    categories: list[str] = []
    for category, _ in EXPENSE_ITEMS:
        if category not in categories:
            categories.append(category)

    section_bar_row = 4
    ws.cell(section_bar_row, 7).value = "BY CATEGORY"
    ws.cell(section_bar_row, 7).font = Font(name=FONT_BODY, size=8.5, bold=True,
                                            color=argb(PALETTE.paper))
    ws.cell(section_bar_row, 8).value = "Monthly"
    ws.cell(section_bar_row, 8).font = Font(name=FONT_BODY, size=8.5, bold=True,
                                            color=argb(PALETTE.paper))
    for column in (7, 8):
        ws.cell(section_bar_row, column).fill = fill(brand.primary)
        ws.cell(section_bar_row, column).alignment = Alignment(
            horizontal="left", vertical="center", indent=1)

    for offset, category in enumerate(categories):
        row = 5 + offset
        name_cell = ws.cell(row, 7)
        name_cell.value = category
        name_cell.font = Font(name=FONT_BODY, size=9.5, color=argb(PALETTE.ink))
        name_cell.fill = fill(PALETTE.paper_warm if offset % 2 == 0 else PALETTE.sand)
        name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        value_cell = ws.cell(row, 8)
        value_cell.value = (f'=SUMIF($A${EXPENSE_FIRST}:$A${EXPENSE_LAST},G{row},'
                            f'$C${EXPENSE_FIRST}:$C${EXPENSE_LAST})')
        value_cell.number_format = MONEY
        value_cell.font = Font(name=FONT_BODY, size=9.5, bold=True,
                               color=argb(PALETTE.gold_dark))
        value_cell.fill = fill(PALETTE.paper_warm if offset % 2 == 0 else PALETTE.sand)
        value_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    roll_total = 5 + len(categories)
    ws.cell(roll_total, 7).value = "TOTAL"
    ws.cell(roll_total, 8).value = f"=SUM(H5:H{roll_total - 1})"
    for column in (7, 8):
        cell = ws.cell(roll_total, column)
        cell.fill = fill(PALETTE.gold_tint)
        cell.font = Font(name=FONT_BODY, size=9.5, bold=True, color=argb(PALETTE.gold_dark))
        cell.number_format = MONEY if column == 8 else "General"
        cell.alignment = Alignment(
            horizontal="right" if column == 8 else "left", vertical="center", indent=1)

    blank_row(ws, EXPENSE_TOTAL + 1, 8)
    note(ws, f"A{EXPENSE_TOTAL + 2}:E{EXPENSE_TOTAL + 4}", "Entering amounts",
         "Enter the average monthly amount for each line. Convert weekly figures by "
         "multiplying by 52 and dividing by 12; convert annual figures by dividing by 12. "
         "Leave a line at zero where it does not apply — the category roll-up on the "
         "right updates automatically.",
         tone="info")

    freeze(ws, "A5")
    print_setup(ws, brand, title_rows="1:4", area=f"A1:H{EXPENSE_TOTAL + 4}")
    ws.sheet_view.showGridLines = False


def build_output(wb: Workbook, brand: BrandProfile) -> None:
    ws = wb.create_sheet(SHEET_OUTPUT)
    widths(ws, {"A": 18, "B": 14, "C": 16, "D": 14, "E": 18, "F": 14,
                "G": 16, "H": 14})

    fact = f"'{SHEET_FACT}'"
    setup = f"'{SHEET_SETUP}'"
    expenses = f"'{SHEET_EXPENSES}'"

    # --- masthead --------------------------------------------------------
    ws.merge_cells("A1:H3")
    head = ws["A1"]
    head.value = f"={setup}!C6"
    head.font = Font(name=FONT_DISPLAY, size=20, bold=True, color=argb(PALETTE.paper))
    head.fill = fill(brand.primary)
    head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for row in range(1, 4):
        ws.row_dimensions[row].height = 22
        for col in range(1, 9):
            ws.cell(row, col).fill = fill(brand.primary)

    ws.merge_cells("A4:H4")
    tagline = ws["A4"]
    tagline.value = f"={setup}!C8"
    tagline.font = Font(name=FONT_BODY, size=9, bold=True, color=argb(PALETTE.obsidian))
    tagline.fill = fill(brand.accent)
    tagline.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, 9):
        ws.cell(4, col).fill = fill(brand.accent)
    ws.row_dimensions[4].height = 18

    ws.merge_cells("A6:H7")
    title = ws["A6"]
    title.value = f"={setup}!F6"
    title.font = Font(name=FONT_DISPLAY, size=15, bold=True, color=argb(PALETTE.ink))
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("A8:H8")
    stamp = ws["A8"]
    stamp.value = (f'={setup}!F7&"  ·  Prepared by "&IF({setup}!F8="","'
                   f'"&{setup}!C6,{setup}!F8)&"  ·  Version "&{setup}!F9')
    stamp.font = Font(name=FONT_BODY, size=9, color=argb(PALETTE.ink_soft))
    stamp.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    blank_row(ws, 9, 8)

    # --- portfolio summary ----------------------------------------------
    section_bar(ws, 10, 8, "PORTFOLIO SUMMARY", brand, kicker="Calculated from the fact find")
    blank_row(ws, 11, 5)
    kpi_tile(ws, 12, 1, 2, "TOTAL ASSETS", f"={fact}!D{ASSET_TOTAL}", brand)
    kpi_tile(ws, 12, 3, 2, "TOTAL DEBT",
             f"={fact}!G{ASSET_TOTAL}+{fact}!F{LIAB_TOTAL}", brand)
    kpi_tile(ws, 12, 5, 2, "NET POSITION", "=A13-C13", brand)
    kpi_tile(ws, 12, 7, 2, "MONTHLY EXPENSES", f"={expenses}!C{EXPENSE_TOTAL}", brand)
    flag_negative(ws, "E13")
    blank_row(ws, 15, 8)

    # --- personal details -------------------------------------------------
    section_bar(ws, 16, 8, "PERSONAL DETAILS", brand)
    for col, text in ((1, "PRIMARY APPLICANT"), (5, "SECONDARY APPLICANT")):
        subhead(ws, 17, col, col + 3, text, brand)

    personal_map = [
        ("Full Name", 'TRIM({f}!C7&" "&{f}!C8&" "&{f}!C9)',
         'TRIM({f}!H7&" "&{f}!H8&" "&{f}!H9)', "General"),
        ("Date of Birth", _blank_safe("{f}!C10"), _blank_safe("{f}!H10"), DATE),
        ("Mobile", _blank_safe("{f}!C15"), _blank_safe("{f}!H15"), "General"),
        ("Email", _blank_safe("{f}!C16"), _blank_safe("{f}!H16"), "General"),
        ("Current Address", _blank_safe("{f}!C19"), _blank_safe("{f}!H19"), "General"),
        ("Living Situation", _blank_safe("{f}!C20"), _blank_safe("{f}!H20"), "General"),
    ]
    for offset, (name, left, right, number_format) in enumerate(personal_map):
        row = 18 + offset
        label(ws, row, 1, name, span=2)
        formula_cell(ws, row, 3, f"={left.format(f=fact)}", span=2,
                     number_format=number_format)
        label(ws, row, 5, name, span=2)
        formula_cell(ws, row, 7, f"={right.format(f=fact)}", span=2,
                     number_format=number_format)
        ws.row_dimensions[row].height = 18
    blank_row(ws, 24, 8)

    # --- employment & income ---------------------------------------------
    section_bar(ws, 25, 8, "EMPLOYMENT & INCOME", brand)
    for col, text in ((1, "PRIMARY APPLICANT"), (5, "SECONDARY APPLICANT")):
        subhead(ws, 26, col, col + 3, text, brand)

    # Row anchors, not guesses: EMPLOYMENT_FIELDS starts at ROW_EMPLOYMENT + 1, so
    # Employment Type is row 27, Employer 28, Role 29, Address 30, Start Date 31.
    employment_map = [
        ("Employer / Business", 28, "General"),
        ("Role / Position", 29, "General"),
        ("Employment Type", 27, "General"),
        ("Start Date", 31, DATE),
    ]
    for offset, (name, source_row, number_format) in enumerate(employment_map):
        row = 27 + offset
        label(ws, row, 1, name, span=2)
        formula_cell(ws, row, 3, f"={_blank_safe(f'{fact}!C{source_row}')}", span=2,
                     number_format=number_format)
        label(ws, row, 5, name, span=2)
        formula_cell(ws, row, 7, f"={_blank_safe(f'{fact}!H{source_row}')}", span=2,
                     number_format=number_format)
        ws.row_dimensions[row].height = 18

    row = 27 + len(employment_map)
    label(ws, row, 1, "Total Annual Income", span=2)
    formula_cell(ws, row, 3, f"=SUM({fact}!C{INCOME_FIRST}:C{INCOME_LAST})", span=2,
                 number_format=MONEY, bold=True, background=PALETTE.gold_tint,
                 colour=PALETTE.gold_dark)
    label(ws, row, 5, "Total Annual Income", span=2)
    formula_cell(ws, row, 7, f"=SUM({fact}!H{INCOME_FIRST}:H{INCOME_LAST})", span=2,
                 number_format=MONEY, bold=True, background=PALETTE.gold_tint,
                 colour=PALETTE.gold_dark)
    ws.row_dimensions[row].height = 18
    blank_row(ws, row + 1, 8)

    # --- position table ---------------------------------------------------
    start = row + 2
    section_bar(ws, start, 8, "ASSETS, LIABILITIES & LIVING EXPENSES", brand,
                kicker="Roll-up — see the fact find for line detail")
    table_header(ws, start + 1, ["Category", "Description", "", "Value / Balance", "",
                                 "Monthly", "", "Source"], brand)
    position = [
        ("Assets", "Total assets recorded in the fact find", f"={fact}!D{ASSET_TOTAL}",
         "", SHEET_FACT),
        ("Liabilities", "Secured loans plus other liabilities",
         f"={fact}!G{ASSET_TOTAL}+{fact}!F{LIAB_TOTAL}",
         f"={fact}!H{ASSET_TOTAL}+{fact}!G{LIAB_TOTAL}", SHEET_FACT),
        ("Living expenses", "Monthly household expenditure", "",
         f"={expenses}!C{EXPENSE_TOTAL}", SHEET_EXPENSES),
        ("Net position", "Assets less liabilities", "=D{net_a}-D{net_l}", "", "Calculated"),
    ]
    table_body(ws, start + 2, start + 5, 8, formats={3: MONEY, 5: MONEY})
    for offset, (category, description, value, monthly, source) in enumerate(position):
        row_index = start + 2 + offset
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=3)
        ws.merge_cells(start_row=row_index, start_column=4, end_row=row_index, end_column=5)
        ws.merge_cells(start_row=row_index, start_column=6, end_row=row_index, end_column=7)
        ws.cell(row_index, 1).value = category
        ws.cell(row_index, 2).value = description
        if value:
            ws.cell(row_index, 4).value = value.format(net_a=start + 2, net_l=start + 3)
        if monthly:
            ws.cell(row_index, 6).value = monthly
        ws.cell(row_index, 8).value = source
        ws.cell(row_index, 4).number_format = MONEY
        ws.cell(row_index, 6).number_format = MONEY
    flag_negative(ws, f"D{start + 5}")

    # --- declaration ------------------------------------------------------
    declaration_row = start + 7
    note(ws, f"A{declaration_row}:H{declaration_row + 3}", "Declaration",
         "The information contained in this fact find has been provided for the purpose "
         "of assessing the client's financial position and coordinating with authorised "
         "finance and property professionals. It does not constitute credit assistance, "
         "financial product advice or legal advice. The client should verify all "
         "information before relying on the generated document.",
         tone="brand")

    sign_row = declaration_row + 5
    for col, span, text in ((1, 4, "Client signature:"), (5, 4, "Date:")):
        ws.merge_cells(start_row=sign_row, start_column=col, end_row=sign_row,
                       end_column=col + span - 1)
        cell = ws.cell(sign_row, col)
        cell.value = f"{text}   " + ("_" * 34 if col == 1 else "____ / ____ / ______")
        cell.font = Font(name=FONT_BODY, size=10, color=argb(PALETTE.ink))
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for offset in range(span):
            ws.cell(sign_row, col + offset).border = Border(
                top=Side(style="thin", color=argb(PALETTE.line_strong)))
    ws.row_dimensions[sign_row].height = 26

    footer_row = sign_row + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    footer = ws.cell(footer_row, 1)
    footer.value = (f'={setup}!C10&"  ·  "&{setup}!C11&"  ·  "&{setup}!C12'
                    f'&"  ·  "&{setup}!C6')
    footer.font = Font(name=FONT_BODY, size=8.5, color=argb(PALETTE.paper))
    footer.fill = fill(brand.primary)
    footer.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, 9):
        ws.cell(footer_row, col).fill = fill(brand.primary)
    ws.row_dimensions[footer_row].height = 20

    disclaimer_row = footer_row + 1
    ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row,
                   end_column=8)
    disclaimer = ws.cell(disclaimer_row, 1)
    disclaimer.value = f"={setup}!F13"
    disclaimer.font = Font(name=FONT_BODY, size=8, color=argb(PALETTE.ink_faint))
    disclaimer.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    print_setup(ws, brand, area=f"A1:H{disclaimer_row}")
    ws.sheet_view.showGridLines = False


def build(brand: BrandProfile, output: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    build_start_here(wb, brand)
    build_setup(wb, brand)
    build_fact_find(wb, brand)
    build_expenses(wb, brand)
    build_output(wb, brand)
    build_lists(wb, brand)

    wb.properties.title = "White-Label Client Fact Find"
    wb.properties.subject = "Client financial position and fact find — white-label template"
    wb.properties.creator = "Aurixa Systems"
    wb.properties.category = "Finance Portal template"
    wb.properties.keywords = "fact find, client form, white label, finance portal, Aurixa"

    wb.active = 0
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "public/templates/finance-portal/Aurixa_White_Label_Client_Fact_Find.xlsx"
    )
    print(build(DEFAULT_BRAND, target))
